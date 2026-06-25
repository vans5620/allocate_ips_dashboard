#!/usr/bin/env python3
"""
Allocate IPS Dashboard - data builder.
Parses the three Portfolio Appraisal exports, classifies every holding into the
8 IPS categories / 6 roofs, joins each account to its assigned model (Client
Repository) and compares actual weights to Model Master targets using a relative
+/-20% band. Writes ips_data.json (consumed by make_html.py).

Usage:
    python3 build_ips.py [IPS_DASHBOARD_FOLDER] [ALLOCATE_FOLDER] [OUT_JSON]

Defaults point at the standard Allocate OneDrive layout.
"""
import openpyxl, json, re, os, sys, glob

# ---------- paths ----------
DASH  = sys.argv[1] if len(sys.argv)>1 else os.getcwd()
ALLOC = sys.argv[2] if len(sys.argv)>2 else os.path.dirname(DASH.rstrip("/"))
OUT   = sys.argv[3] if len(sys.argv)>3 else os.path.join(DASH,"ips_data.json")
AS_OF = os.environ.get("IPS_ASOF","")   # optional override; else read from file

MODEL_MASTER = os.path.join(ALLOC,"02 Model_Master_Allocate IPS.xlsx")
# Repository filename has varied (v2 / no-v2); pick whichever exists.
CLIENT_REPO  = next((p for p in [os.path.join(ALLOC,"03 Client Repository IPS v2.xlsx"),
                                 os.path.join(ALLOC,"03 Client Repository IPS.xlsx")]
                     if os.path.exists(p)),
                    os.path.join(ALLOC,"03 Client Repository IPS v2.xlsx"))

def find_appraisal(scheme):
    """Find an appraisal xlsx in DASH whose name contains the scheme + 'appraisal'."""
    for f in glob.glob(os.path.join(DASH,"*.xlsx")):
        n=os.path.basename(f).lower()
        if scheme.lower() in n and "appraisal" in n:
            return f
    raise FileNotFoundError(f"No appraisal file for {scheme} in {DASH}")

APPR=[(s,find_appraisal(s)) for s in ["Moderate","Aggressive","Equity"]]
# Summit schemes: explicit files (filenames don't contain 'appraisal') - added Jun 2026
for _sname,_sfile in [("Summit Aggressive","Summit Aggressive.xlsx"),
                      ("Summit Equity","Summit Equity.xlsx")]:
    _p=os.path.join(DASH,_sfile)
    if os.path.exists(_p): APPR.append((_sname,_p))
    else: print(f"WARNING: {_sfile} not found - {_sname} skipped")

SECTIONS={"Equity","Fixed Income","Alternative Assets","Cash and Equivalent"}
CATS=["Domestic Equity","International Equity","Stock Basket","Fixed Income",
      "Commodities","InVITs","Liquid ETFs","Cash","MF Applications"]
# Explicit per-scheme cash targets (override Model Master, which implies ~0%)
SCHEME_CASH={"Moderate":0.6,"Aggressive":0.65,"Equity":0.40,
             "Summit Aggressive":0.65,"Summit Equity":0.40}
# Summit -> model sheet override. Summit Aggressive has no own sheet -> Allocate Aggressive.
# (sheet names stored stripped, so 'Summit - Equity ' -> 'Summit - Equity')
SCHEME_MODEL={"Summit Aggressive":"Aggressive V2 - Nov","Summit Equity":"Summit - Equity"}
ROOFS={"Total Equity":["Domestic Equity","International Equity"],
 "Fixed Income":["Fixed Income"],
 "Listed Shares (Stock Basket)":["Stock Basket"],
 "Alternate Assets":["Commodities","InVITs"],
 "Cash":["Cash"],"Liquid":["Liquid ETFs"],"MF Applications":["MF Applications"]}
INTL_KW=["U.S.","US OPP","GREATER CHINA","CHINA","ASIAN","GLOBAL","EMERGING",
         "OVERSEAS","FEEDER","NASDAQ","INTERNATIONAL","S&P 500"]
FUND_MARK=["DIRECT","ETF","BEES"," INDEX"," FUND","FOF","NIFTY 50 VALUE","MOMENTUM 30","LOW VOLATILITY"]

def classify(name, section):
    u=name.upper().strip()
    if "MUTUAL FUND APPLICATION" in u: return "MF Applications"
    if "TAX DEDUCTED" in u or "CASH REC" in u or u=="CASH": return "Cash"
    if section=="Cash and Equivalent": return "Liquid ETFs" if ("LIQUID" in u or "OVERNIGHT" in u) else "Cash"
    if section=="Fixed Income": return "Fixed Income"
    if section=="Alternative Assets": return "InVITs" if "INVIT" in u else "Commodities"
    if section=="Equity":
        if any(k in u for k in INTL_KW): return "International Equity"
        if "NIFTY 500" in u: return "Stock Basket"
        if any(k in u for k in FUND_MARK): return "Domestic Equity"
        return "Stock Basket"
    return "Cash"

def cat_from_model(ac,ins,nm):
    ac=(ac or "").strip(); ins=(ins or "").strip(); u=(nm or "").upper()
    if ac=="Equity":
        if ins=="International Equity": return "International Equity"
        if ins=="Listed": return "Stock Basket"
        if "NIFTY 500" in u: return "Stock Basket"
        return "Domestic Equity"
    if ac=="Fixed Income": return "Fixed Income"
    if ac=="Alternative Assets":
        return "InVITs" if ("INVIT" in ins.upper() or "INVIT" in u) else "Commodities"
    if ac=="Cash and Equivalent":
        return "Liquid ETFs" if ("LIQUID" in ins.upper() or "LIQUID" in u or "OVERNIGHT" in u) else "Cash"
    return None

# ---------- model targets ----------
def load_model_targets(path):
    wbm=openpyxl.load_workbook(path,read_only=True,data_only=True)
    mt={}
    for s in wbm.sheetnames:
        if s.strip()=="Securitymaster": continue
        tgt={c:0.0 for c in CATS}; any_=False
        for r in wbm[s].iter_rows(min_row=2,values_only=True):
            if len(r)<9: continue
            ac,ins,nm,tw=r[1],r[3],r[5],r[8]
            if tw in (None,""): continue
            try: tw=float(tw)
            except: continue
            c=cat_from_model(ac,ins,nm)
            if c: tgt[c]+=tw; any_=True
        if any_:
            tot=sum(tgt.values())
            if tot<=1.5: tgt={k:v*100 for k,v in tgt.items()}
            mt[s.strip()]={c:round(tgt[c],4) for c in CATS}
    wbm.close()
    return mt

model_targets=load_model_targets(MODEL_MASTER)

# scheme -> default model sheet (variant-invariant at bucket level; pick first present)
def scheme_default(scheme):
    if scheme in SCHEME_MODEL and SCHEME_MODEL[scheme] in model_targets:
        return SCHEME_MODEL[scheme]
    for pref in [f"{scheme} V2 - Nov", f"{scheme} - Aug", f"{scheme} - Feb", f"{scheme} - Feb "]:
        if pref.strip() in model_targets: return pref.strip()
    # fallback: any sheet starting with scheme
    for k in model_targets:
        if k.lower().startswith(scheme.lower()): return k
    return None

# ---------- client repository (optional / robust) ----------
repo_by_acct={}; repo_by_code={}
def _s(v): return str(v).strip() if v is not None else ""
try:
    wb=openpyxl.load_workbook(CLIENT_REPO,read_only=True,data_only=True)
    ws=wb["Repository"]
    hdr=[_s(c) for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
    I={h:i for i,h in enumerate(hdr)}
    def col(r,k): return _s(r[I[k]]) if k in I and I[k]<len(r) else ""
    for r in ws.iter_rows(min_row=2,values_only=True):
        if all(c is None for c in r): continue
        model=col(r,"Model Assigned"); rm=col(r,"RM"); nm=col(r,"Client Name")
        info={"model":model,"rm":rm,"name":nm}
        wsid=col(r,"WS Client ID").split(".")[0]
        if wsid: repo_by_acct[wsid]=info
        for k in ("Custody Code","Client Code"):
            code=col(r,k).upper()
            if code: repo_by_code[code]=info
    wb.close()
    print(f"repository loaded: {len(repo_by_acct)} WS IDs, {len(repo_by_code)} codes")
except Exception as e:
    print(f"WARNING: client repository unavailable ({e}). Falling back to scheme-default models.")

# ---------- parse appraisals ----------
def parse(path, scheme):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb.active
    asof=""
    out=[]; cur=None; st=None
    def flush():
        if st and st["acct"]:
            out.append({"acct":st["acct"],"name":st["name"],"code":st["code"],"scheme":scheme,
                        "mv":{c:round(st["bycat"].get(c,0.0),2) for c in CATS},
                        "total":round(st["total"],2)})
    for r in ws.iter_rows(values_only=True):
        a=_s(r[0]); mv=r[5] if len(r)>5 else None; pa=r[8] if len(r)>8 else None
        if a.startswith("As of"): asof=asof or a.replace("As of","").strip()
        if a.startswith("Account :"):
            flush()
            body=a.replace("Account :","").strip()
            m=re.match(r"(\d+)\s+(.*)",body)
            acct=m.group(1) if m else body; rest=m.group(2) if m else ""
            cm=re.search(r"-\s*([A-Za-z]{2,}\d+)\s*$",rest)
            code=cm.group(1).upper() if cm else ""
            name=re.sub(r"\s*-\s*[A-Za-z]{2,}\d+\s*$","",rest).strip()
            st={"acct":acct,"name":name,"code":code,"bycat":{},"total":0.0}; cur=None
            continue
        if a in SECTIONS: cur=a; continue
        if not a or a in ("Security","By Asset Class") or a.startswith(("IONIC","As of","PORTFOLIO","This is")): continue
        if mv in (None,"") or pa in (None,"") or cur is None: continue
        try: mvv=float(mv)
        except: continue
        c=classify(a,cur); st["bycat"][c]=st["bycat"].get(c,0.0)+mvv; st["total"]+=mvv
    flush(); wb.close()
    return out, asof

clients=[]; asof_found=""
for scheme,path in APPR:
    cs,asof=parse(path,scheme); clients+=cs; asof_found=asof_found or asof
AS_OF=AS_OF or asof_found or "—"

# ---------- join + score ----------
def breach(a,t):
    if t is None: return "na",""
    if t==0: return ("compliant","") if a<0.5 else ("breach","over")
    if a<t*0.8: return "breach","under"
    if a>t*1.2: return "breach","over"
    return "compliant",""

out=[]
for c in clients:
    tot=c["total"] or 1.0
    wt={cat:round(c["mv"][cat]/tot*100,4) for cat in CATS}
    # Summit clients are not in the Allocate repository -> always score on scheme basis
    # (also avoids any account-number collision with Allocate WS IDs).
    info = None if c["scheme"].startswith("Summit") else \
           (repo_by_acct.get(c["acct"]) or (repo_by_code.get(c["code"]) if c["code"] else None))
    model_lbl=""; rm="—"; basis=""; tgt=None
    if info:
        rm=info["rm"] or "—"
        assigned=info["model"]
        if assigned in model_targets:
            tgt=model_targets[assigned]; model_lbl=assigned; basis="repository"
        else:
            sd=scheme_default(c["scheme"])
            tgt=model_targets.get(sd); model_lbl=(assigned or c["scheme"])+" (scheme targets)"; basis="scheme"
    else:
        sd=scheme_default(c["scheme"])
        tgt=model_targets.get(sd); model_lbl=c["scheme"]+" (scheme default)"; basis="scheme"
    if tgt is not None and c["scheme"] in SCHEME_CASH:
        tgt=dict(tgt); tgt["Cash"]=SCHEME_CASH[c["scheme"]]
    cat_rows=[]
    for cat in CATS:
        t=tgt[cat] if tgt else None
        sstat,dirn=breach(wt[cat],t)
        cat_rows.append({"cat":cat,"actual":wt[cat],"target":(round(t,2) if t is not None else None),
                         "status":sstat,"dir":dirn,"mv":c["mv"][cat]})
    roof_rows=[]
    for roof,cats in ROOFS.items():
        a=round(sum(wt[x] for x in cats),4)
        t=round(sum(tgt[x] for x in cats),4) if tgt else None
        sstat,dirn=breach(a,t)
        roof_rows.append({"roof":roof,"actual":a,"target":(round(t,2) if t is not None else None),
                          "status":sstat,"dir":dirn})
    out.append({"acct":c["acct"],"name":(info["name"] if info and info["name"] else c["name"]),
                "scheme":c["scheme"],"model":model_lbl,"rm":rm,"basis":basis,
                "mapped":tgt is not None,"total":c["total"],"roofs":roof_rows,"cats":cat_rows})

data={"as_of":AS_OF,"cats":CATS,"roofs":list(ROOFS.keys()),"clients":out,
      "n_total":len(out),"n_mapped":sum(1 for x in out if x["mapped"]),
      "n_repo":sum(1 for x in out if x["basis"]=="repository")}
json.dump(data,open(OUT,"w"))
print(f"clients: {data['n_total']} | scored: {data['n_mapped']} | exact-repo: {data['n_repo']} | as of {AS_OF}")
bad=sum(1 for x in out if abs(sum(r['actual'] for r in x['cats'])-100)>0.5)
print("category-weight sums off >0.5pp:",bad)
